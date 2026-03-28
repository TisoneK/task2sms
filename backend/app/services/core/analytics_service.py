from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, case
from datetime import datetime, timezone, timedelta
from typing import Optional
from app.models.task import Task, TaskStatus
from app.models.notification import Notification, NotificationStatus
from app.models.email_notification import EmailNotification, EmailStatus
from app.models.whatsapp import WhatsAppMessage, WhatsAppStatus
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


async def get_full_analytics(
    db: AsyncSession,
    user_id: int,
    days: int = 30,
) -> dict:
    now = datetime.now(timezone.utc)
    since = now - timedelta(days=days)

    # ── SMS ────────────────────────────────────────────────────────
    sms_total = await db.execute(
        select(func.count(Notification.id))
        .where(Notification.user_id == user_id, Notification.created_at >= since)
    )
    sms_sent = await db.execute(
        select(func.count(Notification.id))
        .where(Notification.user_id == user_id,
               Notification.status == NotificationStatus.SENT,
               Notification.created_at >= since)
    )
    sms_failed = await db.execute(
        select(func.count(Notification.id))
        .where(Notification.user_id == user_id,
               Notification.status == NotificationStatus.FAILED,
               Notification.created_at >= since)
    )

    # ── Daily SMS breakdown ────────────────────────────────────────
    daily_rows = await db.execute(
        select(
            func.date(Notification.created_at).label("day"),
            func.count(Notification.id).label("total"),
            func.sum(case((Notification.status == NotificationStatus.SENT, 1), else_=0)).label("sent"),
            func.sum(case((Notification.status == NotificationStatus.FAILED, 1), else_=0)).label("failed"),
        )
        .where(Notification.user_id == user_id, Notification.created_at >= since)
        .group_by(func.date(Notification.created_at))
        .order_by(func.date(Notification.created_at))
    )
    daily_sms = [{"day": str(r.day), "total": r.total,
                  "sent": r.sent or 0, "failed": r.failed or 0}
                 for r in daily_rows]

    # ── Provider breakdown ─────────────────────────────────────────
    provider_rows = await db.execute(
        select(Notification.provider, func.count(Notification.id))
        .where(Notification.user_id == user_id, Notification.created_at >= since)
        .group_by(Notification.provider)
    )
    providers = {p: c for p, c in provider_rows}

    # ── Email ──────────────────────────────────────────────────────
    email_total = await db.execute(
        select(func.count(EmailNotification.id))
        .where(EmailNotification.user_id == user_id, EmailNotification.created_at >= since)
    )
    email_sent = await db.execute(
        select(func.count(EmailNotification.id))
        .where(EmailNotification.user_id == user_id,
               EmailNotification.status == EmailStatus.SENT,
               EmailNotification.created_at >= since)
    )

    # ── WhatsApp ───────────────────────────────────────────────────
    wa_total = await db.execute(
        select(func.count(WhatsAppMessage.id))
        .where(WhatsAppMessage.user_id == user_id, WhatsAppMessage.created_at >= since)
    )
    wa_sent = await db.execute(
        select(func.count(WhatsAppMessage.id))
        .where(WhatsAppMessage.user_id == user_id,
               WhatsAppMessage.status == WhatsAppStatus.SENT,
               WhatsAppMessage.created_at >= since)
    )

    # ── Tasks ──────────────────────────────────────────────────────
    task_counts = await db.execute(
        select(Task.status, func.count(Task.id))
        .where(Task.user_id == user_id)
        .group_by(Task.status)
    )
    task_map = {str(s): c for s, c in task_counts}

    # ── Top tasks by runs ──────────────────────────────────────────
    top_tasks_rows = await db.execute(
        select(Task.name, Task.run_count, Task.fail_count)
        .where(Task.user_id == user_id)
        .order_by(Task.run_count.desc())
        .limit(5)
    )
    top_tasks = [{"name": n, "runs": r, "fails": f} for n, r, f in top_tasks_rows]

    t = sms_total.scalar() or 0
    s = sms_sent.scalar() or 0

    return {
        "period_days": days,
        "sms": {
            "total": t,
            "sent": s,
            "failed": sms_failed.scalar() or 0,
            "delivery_rate": round(s / t * 100, 1) if t else 0,
        },
        "email": {
            "total": email_total.scalar() or 0,
            "sent": email_sent.scalar() or 0,
        },
        "whatsapp": {
            "total": wa_total.scalar() or 0,
            "sent": wa_sent.scalar() or 0,
        },
        "tasks": {
            "total": sum(task_map.values()),
            **{k: task_map.get(k, 0) for k in ["active", "paused", "failed", "completed"]},
        },
        "daily_sms": daily_sms,
        "providers": providers,
        "top_tasks": top_tasks,
    }


async def export_notifications_xlsx(db: AsyncSession, user_id: int) -> bytes:
    """Generate Excel report of all notifications."""
    result = await db.execute(
        select(Notification)
        .where(Notification.user_id == user_id)
        .order_by(Notification.created_at.desc())
        .limit(10000)
    )
    notifications = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SMS Notifications"

    header_fill = PatternFill("solid", fgColor="0284C7")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["ID", "Recipient", "Message", "Provider", "Status",
               "Message ID", "Error", "Retry Count", "Sent At", "Created At"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 20

    for row_idx, n in enumerate(notifications, 2):
        ws.append([
            n.id, n.recipient, n.message[:200], n.provider,
            str(n.status.value), n.provider_message_id or "",
            n.error_message or "", n.retry_count,
            str(n.sent_at)[:19] if n.sent_at else "",
            str(n.created_at)[:19] if n.created_at else "",
        ])

    # auto-width
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
